# https://github.com/RapidAI/TableStructureRec
# 安装：pip install wired_table_rec lineless_table_rec table_cls
# https://github.com/RapidAI/TableStructureRec
# 安装：pip install wired_table_rec lineless_table_rec table_cls

import os
import cv2
import openpyxl
# from openpyxl.styles import PatternFill
import pandas as pd
from table_cls import TableCls
from wired_table_rec.main import WiredTableInput, WiredTableRecognition
from lineless_table_rec.main import LinelessTableInput, LinelessTableRecognition
from rapidocr import RapidOCR
import yaml
from utils import resource_path, rgb_to_named_color

config_path=resource_path("configs/config.yaml")


with open(config_path, "r", encoding="utf-8") as file:
    config = yaml.safe_load(file)


def extract_cell_color(img, bbox):
    """提取单元格区域的平均颜色"""

    # 解析四个角点
    x_coords = bbox[0::2]  # 取偶数索引（x 坐标）
    y_coords = bbox[1::2]  # 取奇数索引（y 坐标）

    x1, x2 = int(min(x_coords)), int(max(x_coords))
    y1, y2 = int(min(y_coords)), int(max(y_coords))

    # 裁剪单元格区域
    cell_region = img[y1:y2, x1:x2]

    # 计算平均颜色（BGR）
    avg_color = cv2.mean(cell_region)[:3]
    return avg_color


def rgb_to_hex(color):
    """将 BGR 颜色转换为 HEX 颜色"""
    b, g, r = map(lambda x: int(round(x)), color)  # BGR 转 RGB，四舍五入后转换为整数
    return f"{r:02X}{g:02X}{b:02X}"


def add_color_to_text(img_path, df, cell_bboxes, output_path):

    img = cv2.imread(img_path)

    #  创建 Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    # 找到第二行对应的 index
    index = next(
        (
            index + 1
            for index, cell_bboxe in enumerate(cell_bboxes)
            if cell_bboxe[1] != cell_bboxes[index + 1][1]
        ),
        None,
    )
    for row_idx, row in df.iterrows():  # 遍历 DataFrame 的每一行
        for col_idx, cell_value in enumerate(row):  # 遍历每一列
            if config["use_team_img"] == True:
                if row_idx == 0:
                    continue

                bbox_index = (
                    (row_idx - 1) * len(row) + index + col_idx
                )  # 计算 bbox 的索引
            else:
                bbox_index = row_idx * len(row) + col_idx
            # 检查索引是否超出范围
            if bbox_index >= len(cell_bboxes):
                print(
                    f"Warning: bbox index {bbox_index} out of bounds for cell_bboxes with size {len(cell_bboxes)}"
                )
                continue
            bbox = cell_bboxes[bbox_index]  # 获取对应单元格的 bbox

            # 解析 bbox 坐标
            try:
                avg_color = extract_cell_color(img, bbox)  # 计算颜色
            except Exception as e:
                print(f"Error processing bbox {bbox}: {e}")
                continue
            b, g, r = map(lambda x: int(round(x)), avg_color)  # BGR 转 RGB
            color = (r, g, b)  # RGB 颜色
            named_color = rgb_to_named_color(color)  # 转换为主要颜色名称

            # 将颜色名称添加到文字前
            cell_text = (
                f"{named_color}{cell_value}"
                if pd.notna(cell_value)
                else f"{named_color}"
            )

            excel_cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=cell_text)

            # #填充表格颜色
            # hex_color = rgb_to_hex(avg_color)  # 转换为 HEX 颜色
            # excel_cell.fill = PatternFill(
            #     start_color=hex_color, end_color=hex_color, fill_type="solid"
            # )

    # 7. 保存 Excel
    wb.save(output_path)
    print(f"Excel 生成成功: {output_path}")


def image_to_excel(img_path, output_path):

    # Init
    wired_input = WiredTableInput()
    lineless_input = LinelessTableInput()
    wired_engine = WiredTableRecognition(wired_input)
    lineless_engine = LinelessTableRecognition(lineless_input)
    # 小yolo模型(0.1s)：TableCls();精度更高的yolox(0.25s)：TableCls(model_type="yolox");更快的qanything(0.07s)模型TableCls(model_type="q")或paddle模型(0.03s)
    table_cls = TableCls(model_type="yolox")
    cls, elasp = table_cls(img_path)
    if cls == "wired":
        table_engine = wired_engine
    else:
        table_engine = lineless_engine

    # 使用RapidOCR输入
    ocr_engine = RapidOCR()
    rapid_ocr_output = ocr_engine(img_path, return_word_box=True)

    # 1.使用RapidOCR输入
    ocr_result = list(
        zip(rapid_ocr_output.boxes, rapid_ocr_output.txts, rapid_ocr_output.scores)
    )
    # # 2.使用单字识别
    # word_results = rapid_ocr_output.word_results
    # ocr_result = [
    #     [word_result[2], word_result[0], word_result[1]] for word_result in word_results
    # ]

    table_results = table_engine(img_path, ocr_result=ocr_result)
    pred_html = table_results.pred_html
    dfs = pd.read_html(pred_html)
    if dfs:
        df = dfs[0]
    else:
        print("未找到 HTML 表格，无法保存为 .xlsx 文件")
    if config["use_color"] == False:
        df.to_excel(output_path, index=False, engine="openpyxl")
        print(f".xlsx 文件已保存到: {output_path}")
    else:
        if hasattr(table_results, "cell_bboxes"):
            cell_bboxes = table_results.cell_bboxes  # 获取单元格坐标
        else:
            raise ValueError("table_results 没有 cell_bboxes 属性，无法解析表格")
        add_color_to_text(img_path, df, cell_bboxes, output_path)


def main():
    img_dir = resource_path("images")
    img_files = [
        f for f in os.listdir(img_dir) if f.lower().endswith((".jpg", ".png", ".jpeg"))
    ]

    if not img_files:
        raise FileNotFoundError("outputs 文件夹中未找到任何 image 文件")

    img_path = os.path.join(img_dir, img_files[0])

    output_path = resource_path(os.path.join("outputs", f"{config['excel_name']}.xlsx"))

    image_to_excel(img_path, output_path)


if __name__ == "__main__":
    main()
